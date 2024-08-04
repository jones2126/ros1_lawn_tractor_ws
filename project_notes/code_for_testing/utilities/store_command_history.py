#!/usr/bin/env python3
'''
Script that looks stores command history in a .xlsx file along with its date.

$ python3 ~/ros1_lawn_tractor_ws/project_notes/code_for_testing/utilities/store_command_history.py

'''
print("Starting store_command_history.py")
import os
import time
import pandas as pd
from openpyxl import load_workbook

# Define paths
history_file = os.path.expanduser("~/.bash_history")
xlsx_file = os.path.expanduser("~/command_line_history.xlsx")
temp_file = os.path.expanduser("~/command_line_history_temp.txt")

# Function to convert epoch to human-readable date
def epoch_to_human(epoch_time):
    return time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(epoch_time))

# Read the existing history file
print("Reading history file")
with open(history_file, 'r') as file:
    lines = file.readlines()

# Prepare data for DataFrame
print("putting history file in DataFrame")
data = []
i = 0
while i < len(lines):
    if lines[i].startswith('#'):
        epoch_time = int(lines[i][1:].strip())
        human_time = epoch_to_human(epoch_time)
        command = lines[i+1].strip()
        data.append([epoch_time, human_time, command])
        i += 2
    else:
        i += 1

# Load existing data from xlsx file if it exists
if os.path.exists(xlsx_file):
    df_existing = pd.read_excel(xlsx_file)
    last_epoch_time = df_existing['Epoch Time'].max()
else:
    df_existing = pd.DataFrame(columns=['Epoch Time', 'Human Readable Time', 'Command'])
    last_epoch_time = 0

# Filter out already stored commands
print("Filtering out already existing commands")
new_data = [row for row in data if row[0] > last_epoch_time]

# Append new data to the DataFrame
print("appending new data to DataFrame")
df_new = pd.DataFrame(new_data, columns=['Epoch Time', 'Human Readable Time', 'Command'])
df_combined = pd.concat([df_existing, df_new], ignore_index=True)

# Save to Excel file
print("Saving data to .xlsx file")
df_combined.to_excel(xlsx_file, index=False)

# Set column width and alignment
print("Adjusting column width and alignment")
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

wb = load_workbook(xlsx_file)
ws = wb.active

column_letter = get_column_letter(1)  # 1 corresponds to column A
ws.column_dimensions[column_letter].width = 13  # 13 characters
for cell in ws[column_letter]:
    cell.alignment = Alignment(horizontal='right')
column_letter = get_column_letter(2)  # 2 corresponds to column B
ws.column_dimensions[column_letter].width = 23  # 23 characters
for cell in ws[column_letter]:
    cell.alignment = Alignment(horizontal='right')
column_letter = get_column_letter(3)  # 3 corresponds to column C
ws.column_dimensions[column_letter].width = 100  # 100 characters
print("Saving file")
wb.save(xlsx_file)

# Print a message indicating completion
print(f"History updated. {len(new_data)} new commands added.")

