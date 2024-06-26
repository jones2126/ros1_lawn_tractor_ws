#!/usr/bin/env python

'''
Read the x, y data for the path, calculate the angle and add the lookahead and speed data.

$ python3 /home/tractor/ros1_lawn_tractor_ws/project_notes/code_for_testing/archive/path_polygon_rings/path_add_angle_to_xy_for_outer_ring.py
'''
import openpyxl
import pandas as pd
import math

def update_df_with_angle_lookahead_speed(input_file_path, input_sheet_name, lookahead, speed, output_sheet_name):
    def calculate_angle(x1, y1, x2, y2):
        '''Calculates the directional angle with respect to the positive X-axis.'''
        angle = math.atan2(y2 - y1, x2 - x1)
        if angle < 0:
            angle += 2 * math.pi
        return angle

    print("Calculating angles")
    df = pd.read_excel(input_file_path, sheet_name=input_sheet_name, header=0)

    # Ensure the DataFrame has the expected columns
    if 'X' not in df.columns or 'Y' not in df.columns:
        raise ValueError("Input sheet does not contain the required columns 'X' and 'Y'")

    # Initialize the angles list
    angles = []
    angles.append(0.0)  # First row gets an angle of 0.0
    for idx in range(1, len(df), 2):
        x1, y1 = df.iloc[idx]['X'], df.iloc[idx]['Y']

        # Check if the next index is within the bounds of the DataFrame
        if idx + 1 < len(df):
            x2, y2 = df.iloc[idx + 1]['X'], df.iloc[idx + 1]['Y']
            angle = calculate_angle(x1, y1, x2, y2)
            angles.extend([angle, angle])

    # Ensure that the length of angles matches the DataFrame
    while len(angles) < len(df):
        angles.append(angles[-1])

    print(f"Length of angles list {len(angles)}, length of DataFrame {len(df)}")
    print("Adding lookahead and speed")
    df['theta'] = angles
    df['lookahead'] = lookahead
    df['speed'] = speed

    # Rename columns to match the required output format
    df.rename(columns={'X': 'x', 'Y': 'y'}, inplace=True)
    df = df[['x', 'y', 'theta', 'lookahead', 'speed']]

    print("Writing to Excel file")
    workbook = openpyxl.load_workbook(input_file_path)

    # Check if the sheet already exists and delete it if it does
    if output_sheet_name in workbook.sheetnames:
        del workbook[output_sheet_name]

    # Create a new sheet
    output_sheet = workbook.create_sheet(output_sheet_name)

    # Write the DataFrame to the new sheet
    for row in df.itertuples(index=False, name=None):
        output_sheet.append(row)

    workbook.save(input_file_path)
    print(f"Data published to sheet: {output_sheet_name} in file: {input_file_path}")

input_file_path = '/home/tractor/ros1_lawn_tractor_ws/project_notes/paths/Collins_Dr_62/site1_20240513/collins_dr_62_A_from_rosbag_step1_20240513_2.xlsx'
input_sheet_name = 'UpdatedPath'
output_sheet_name = 'outer_rings_path'
lookahead = 2.5
speed = 0.75

update_df_with_angle_lookahead_speed(input_file_path, input_sheet_name, lookahead, speed, output_sheet_name)
