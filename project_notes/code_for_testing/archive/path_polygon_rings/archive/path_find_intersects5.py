#!/usr/bin/env python

'''
Script to find intersections between a robot's path and a circular obstacle, and modify the path to avoid the obstacle.

$ python3 /home/tractor/ros1_lawn_tractor_ws/project_notes/code_for_testing/archive/path_polygon_rings/path_find_intersects5.py
'''

print("path_find_intersects5.py starting....")

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import sys

# Function to get ring starting positions
def get_ring_starting_positions(xlsx_file_path):
    df = pd.read_excel(xlsx_file_path, sheet_name='RawInnerRings', engine='openpyxl')
    path_index_changes = df[df['Path_Index'].diff() != 0].reset_index(drop=True)
    ring_starting_positions = [(df.iloc[0]['X'], df.iloc[0]['Y'])] + list(zip(path_index_changes['X'], path_index_changes['Y']))
    
    # Remove duplicates if the first position is repeated
    if ring_starting_positions[0] == ring_starting_positions[1]:
        ring_starting_positions.pop(1)
    
    return ring_starting_positions

# Function to update Path_Index based on starting positions
def update_path_index_based_on_starting_positions(xlsx_file_path, ring_starting_positions):
    try:
        df = pd.read_excel(xlsx_file_path, sheet_name='NewRingPath', engine='openpyxl')
    except ValueError:
        df = pd.DataFrame(columns=['X', 'Y', 'Path_Index'])
    
    path_index = 0

    for i in range(len(df)):
        x = df.at[i, 'X']
        y = df.at[i, 'Y']

        # Check if the current position matches any of the starting positions
        for j, (start_x, start_y) in enumerate(ring_starting_positions):
            if np.isclose(x, start_x) and np.isclose(y, start_y):
                path_index = j
                break
        
        df.at[i, 'Path_Index'] = path_index
    
    return df

# Function to check if two segments intersect
def do_segments_intersect(seg1_start, seg1_end, seg2_start, seg2_end):
    def cross_product(a, b):
        return a[0] * b[1] - a[1] * b[0]

    def subtract_vectors(a, b):
        return (a[0] - b[0], a[1] - b[1])

    def is_point_on_segment(seg_start, seg_end, point):
        seg = subtract_vectors(seg_end, seg_start)
        pt = subtract_vectors(point, seg_start)
        cross_prod = cross_product(seg, pt)
        dot_prod = seg[0] * pt[0] + seg[1] * pt[1]
        squared_len_seg = seg[0] ** 2 + seg[1] ** 2
        return abs(cross_prod) < 1e-7 and 0 <= dot_prod <= squared_len_seg

    seg1_vec = subtract_vectors(seg1_end, seg1_start)
    seg2_vec = subtract_vectors(seg2_end, seg2_start)
    seg_start_diff = subtract_vectors(seg2_start, seg1_start)
    seg_end_diff = subtract_vectors(seg2_end, seg1_start)

    seg1_cross_start = cross_product(seg1_vec, seg_start_diff)
    seg1_cross_end = cross_product(seg1_vec, seg_end_diff)

    seg2_cross_start = cross_product(seg2_vec, seg_start_diff)
    seg2_cross_end = cross_product(seg2_vec, subtract_vectors(seg1_end, seg2_start))

    if seg1_cross_start * seg1_cross_end <= 0 and seg2_cross_start * seg2_cross_end <= 0:
        if seg1_cross_start == 0 and is_point_on_segment(seg1_start, seg1_end, seg2_start):
            return True
        if seg1_cross_end == 0 and is_point_on_segment(seg1_start, seg1_end, seg2_end):
            return True
        if seg2_cross_start == 0 and is_point_on_segment(seg2_start, seg2_end, seg1_start):
            return True
        if seg2_cross_end == 0 and is_point_on_segment(seg2_start, seg2_end, seg1_end):
            return True
        return True
    return False

# Function to find intersections with circle
def find_intersections_with_circle(xlsx_file_path, df_raw_inner_rings):
    try:
        df_new_ring_path = pd.read_excel(xlsx_file_path, sheet_name='NewRingPath', engine='openpyxl')
        print(f"Read NewRingPath sheet: {df_new_ring_path}")
    except ValueError:
        df_new_ring_path = df_raw_inner_rings[['X', 'Y', 'Path_Index']].copy()
        print(f"NewRingPath sheet does not exist. Using data from RawInnerRings.")

    df_obstacle_circle = pd.read_excel(xlsx_file_path, sheet_name='Obstacle', engine='openpyxl')
    print(f"Read Obstacle sheet: {df_obstacle_circle}")

    circle_center = (df_obstacle_circle.iloc[0]['X'], df_obstacle_circle.iloc[0]['Y'])
    circle_radius = df_obstacle_circle.iloc[0]['Radius']

    print(f"Circle center: {circle_center}, Circle radius: {circle_radius}")

    def is_point_in_circle(point, center, radius):
        return np.sqrt((point[0] - center[0])**2 + (point[1] - center[1])**2) <= radius

    def find_intersection_point(seg_start, seg_end, center, radius):
        d = np.subtract(seg_end, seg_start)
        f = np.subtract(seg_start, center)
        a = np.dot(d, d)
        b = 2 * np.dot(f, d)
        c = np.dot(f, f) - radius**2
        discriminant = b**2 - 4*a*c
        
        if discriminant < 0:
            return None
        else:
            discriminant = np.sqrt(discriminant)
            t1 = (-b - discriminant) / (2*a)
            t2 = (-b + discriminant) / (2*a)
            intersection_points = []
            if 0 <= t1 <= 1:
                intersection_points.append(seg_start + t1 * d)
            if 0 <= t2 <= 1:
                intersection_points.append(seg_start + t2 * d)
            return intersection_points if intersection_points else None

    revised_path = pd.DataFrame(columns=['X', 'Y', 'Path_Index'])
    intersection_points = []
    intersecting_segments = []
    intersecting_index = []

    if len(df_new_ring_path) > 5000:
        print("Program stopped due to too many records in df_new_ring_path.")
        sys.exit()    

    for i in range(len(df_new_ring_path) - 1):
        seg_start = np.array([df_new_ring_path.at[i, 'X'], df_new_ring_path.at[i, 'Y']])
        seg_end = np.array([df_new_ring_path.at[i + 1, 'X'], df_new_ring_path.at[i + 1, 'Y']])
        
        print(f"Checking segment {i} from {seg_start} to {seg_end}")

        if is_point_in_circle(seg_start, circle_center, circle_radius) or is_point_in_circle(seg_end, circle_center, circle_radius):
            print(f"Segment {i} intersects with circle")
            intersection_points_list = find_intersection_point(seg_start, seg_end, circle_center, circle_radius)
            if intersection_points_list is not None:
                for intersection_point in intersection_points_list:
                    print(f"Intersection found at {intersection_point}")
                    intersection_points.append(intersection_point)
                    intersecting_segments.append((seg_start, seg_end))
                    intersecting_index.append(i)
        else:
            print(f"Segment {i} does not intersect with circle")

    if intersection_points:
        for i in range(len(intersecting_index)):
            segment_index = intersecting_index[i]
            segment_start, segment_end = intersecting_segments[i]
            intersection_point = intersection_points[i]

            temp_path = df_new_ring_path.iloc[:segment_index + 1].copy()
            temp_path = pd.concat([temp_path, pd.DataFrame({'X': [intersection_point[0]], 'Y': [intersection_point[1]], 'Path_Index': [df_new_ring_path.at[segment_index, 'Path_Index']]})], ignore_index=True)
            
            revised_path = pd.concat([revised_path, temp_path], ignore_index=True)
        
        # Append the remaining path after the last intersection
        revised_path = pd.concat([revised_path, df_new_ring_path.iloc[segment_index + 1:].copy()], ignore_index=True)
    else:
        print("No intersections found, copying original path.")
        revised_path = df_new_ring_path

    print(f"Revised path:\n{revised_path}")

    # Check the length of the revised path
    if len(revised_path) > 5000:
        print(f"Error: Revised path has too many records ({len(revised_path)}). Stopping program.")
        return None, None

    # Create a DataFrame to store debug data
    debug_data = pd.DataFrame({
        'Intersection_X': [point[0] for point in intersection_points],
        'Intersection_Y': [point[1] for point in intersection_points],
        'Segment_Start_X': [seg[0][0] for seg in intersecting_segments],
        'Segment_Start_Y': [seg[0][1] for seg in intersecting_segments],
        'Segment_End_X': [seg[1][0] for seg in intersecting_segments],
        'Segment_End_Y': [seg[1][1] for seg in intersecting_segments],
        'Path_Index': intersecting_index
    })

    # Add the new sheet 'NewRingPath' to the .xlsx file, removing it first if it already exists
    with pd.ExcelWriter(xlsx_file_path, engine='openpyxl', mode='a') as writer:
        workbook = writer.book
        if 'NewRingPath' in workbook.sheetnames:
            del workbook['NewRingPath']
        revised_path.to_excel(writer, sheet_name='NewRingPath', index=False)
        
        # Write the debug data to a new sheet
        debug_data.to_excel(writer, sheet_name='DebugData', index=False)

    return revised_path, debug_data

# Main Function
xlsx_file_path = '/home/tractor/ros1_lawn_tractor_ws/project_notes/paths/Collins_Dr_62/site1_20240513/collins_dr_62_A_from_rosbag_step1_20240513_2.xlsx'

# Adding debug to print initial path data
df_raw_inner_rings = pd.read_excel(xlsx_file_path, sheet_name='RawInnerRings', engine='openpyxl')
print(f"Initial data from RawInnerRings:\n{df_raw_inner_rings}")

revised_path, debug_data = find_intersections_with_circle(xlsx_file_path, df_raw_inner_rings)

if revised_path is None:
    print("Program stopped due to too many records in the revised path.")
    sys.exit()

print("Getting starting positions")
ring_starting_positions = get_ring_starting_positions(xlsx_file_path)
print("Starting positions:", ring_starting_positions)
print("Updating Path_Index")
updated_new_ring_path = update_path_index_based_on_starting_positions(xlsx_file_path, ring_starting_positions)

# Check the length of the updated new ring path before writing to the file
if len(updated_new_ring_path) > 5000:
    print(f"Error: Updated new ring path has too many records ({len(updated_new_ring_path)}). Stopping program.")
    sys.exit()

# Add the updated 'NewRingPath' sheet to the .xlsx file, removing it first if it already exists
with pd.ExcelWriter(xlsx_file_path, engine='openpyxl', mode='a') as writer:
    workbook = writer.book
    if 'NewRingPath' in workbook.sheetnames:
        del workbook['NewRingPath']
    updated_new_ring_path.to_excel(writer, sheet_name='NewRingPath', index=False)

print("Path_Index updated and saved to 'NewRingPath' sheet.")

# Read the updated 'NewRingPath' for plotting
df_new_ring_path = pd.read_excel(xlsx_file_path, sheet_name='NewRingPath', engine='openpyxl')

# Plot the values
plt.figure(figsize=(10, 6))
for path_index in df_new_ring_path['Path_Index'].unique():
    path_data = df_new_ring_path[df_new_ring_path['Path_Index'] == path_index]
    plt.scatter(path_data['X'], path_data['Y'], label=f'Path Index {path_index}', s=10)

# Plotting the obstacle circle for reference
df_obstacle_circle = pd.read_excel(xlsx_file_path, sheet_name='Obstacle', engine='openpyxl')
circle_center = (df_obstacle_circle.iloc[0]['X'], df_obstacle_circle.iloc[0]['Y'])
circle_radius = df_obstacle_circle.iloc[0]['Radius']
circle = plt.Circle(circle_center, circle_radius, color='green', fill=False, label='Obstacle Circle')
plt.gca().add_patch(circle)

# Customize the plot
plt.xlabel('X')
plt.ylabel('Y')
plt.legend()
plt.title('New Ring Path')
plt.grid(True)
plt.axis('equal')

# Show the plot
plt.show()
