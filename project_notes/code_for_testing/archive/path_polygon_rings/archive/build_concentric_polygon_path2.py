#!/usr/bin/env python

'''
Script to find intersections between circles defined in 'Obstcl_list' and the line segments in 'RawInnerRings'.

$ python3 /home/tractor/ros1_lawn_tractor_ws/project_notes/code_for_testing/archive/path_polygon_rings/build_concentric_polygon_path2.py
'''
print("Running build_concentric_polygon_path2.py")


import pandas as pd
from openpyxl import load_workbook

# Define the path to your Excel file
xlsx_file_path = '/home/tractor/ros1_lawn_tractor_ws/project_notes/paths/Collins_Dr_62/site1_20240513/collins_dr_62_A_from_rosbag_step1_20240513_2.xlsx'

# Load data from Excel sheets
raw_inner_rings = pd.read_excel(xlsx_file_path, sheet_name='RawInnerRings')
intersection_points = pd.read_excel(xlsx_file_path, sheet_name='IntersectionPoints')
arc_path = pd.read_excel(xlsx_file_path, sheet_name='ArcPath')

# Initialize an empty DataFrame for the new waypoints
new_waypoints = pd.DataFrame(columns=raw_inner_rings.columns)

# Process each Path_Index in RawInnerRings
unique_path_indices = raw_inner_rings['Path_Index'].unique()

for path_index in unique_path_indices:
    # Extract waypoints for the current Path_Index
    path_data = raw_inner_rings[raw_inner_rings['Path_Index'] == path_index]
    
    # Get corresponding intersection points for the current Path_Index
    intersection_data = intersection_points[intersection_points['Path_Index'] == path_index]
    row_indices = intersection_data['Row_Index'].values
    
    last_index = 0
    
    for i in range(len(row_indices) - 1):
        valid_start = last_index
        valid_end = row_indices[i] - 1
        
        # Append valid waypoints to the new list
        new_waypoints = pd.concat([new_waypoints, path_data.iloc[valid_start:valid_end+1]], ignore_index=True)
        
        replacement_start = row_indices[i]
        replacement_end = row_indices[i + 1] - 1
        
        # Find corresponding replacement waypoints from ArcPath
        ref_value = intersection_data.iloc[i]['Reference']
        replacement_data = arc_path[(arc_path['Path_Index'] == path_index) & (arc_path['Reference'] == ref_value)]
        
        # Append replacement waypoints, ensuring Path_Index matches
        new_waypoints = pd.concat([new_waypoints, replacement_data], ignore_index=True)
        
        last_index = replacement_end + 1
    
    # Append remaining waypoints
    remaining_waypoints = path_data.iloc[last_index:]
    new_waypoints = pd.concat([new_waypoints, remaining_waypoints], ignore_index=True)

# Load the workbook
book = load_workbook(xlsx_file_path)

# Check if the 'UpdatedPath' sheet exists and delete it if it does
if 'UpdatedPath' in book.sheetnames:
    del book['UpdatedPath']
    book.save(xlsx_file_path)

# Save the new waypoints to a new sheet in the existing Excel file
with pd.ExcelWriter(xlsx_file_path, engine='openpyxl', mode='a') as writer:
    new_waypoints.to_excel(writer, sheet_name='UpdatedPath', index=False)

print(f"Updated waypoints saved to a new sheet 'UpdatedPath' in {xlsx_file_path}")

# Prompt the user for further action
action = input("The initial path has been created, (S)top for review or (C)ontinue: ").strip().lower()

if action == 'c':
    # Load the newly created sheet
    updated_path = pd.read_excel(xlsx_file_path, sheet_name='UpdatedPath')
    
    # Print the columns to debug the issue
    print("Columns in 'UpdatedPath':", updated_path.columns)
    
    # Check rows where columns Arc_X and Arc_Y have data
    if 'Arc_X' in updated_path.columns and 'Arc_Y' in updated_path.columns:
        mask = updated_path['Arc_X'].notnull() & updated_path['Arc_Y'].notnull()
        
        # Copy data from columns Arc_X and Arc_Y to X and Y where Arc_X and Arc_Y are not null
        updated_path.loc[mask, 'X'] = updated_path.loc[mask, 'Arc_X']
        updated_path.loc[mask, 'Y'] = updated_path.loc[mask, 'Arc_Y']
        
        # Drop columns Arc_X and Arc_Y
        updated_path.drop(columns=['Arc_X', 'Arc_Y'], inplace=True)
        
        # Reorder columns if necessary
        updated_path = updated_path[['Path_Index', 'X', 'Y'] + [col for col in updated_path.columns if col not in ['Path_Index', 'X', 'Y']]]
        
        # Save the updated DataFrame back to the same sheet
        with pd.ExcelWriter(xlsx_file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            updated_path.to_excel(writer, sheet_name='UpdatedPath', index=False)

        print(f"Columns Arc_X and Arc_Y moved to X and Y where not empty, and saved back to 'UpdatedPath' in {xlsx_file_path}")
    else:
        print("Columns 'Arc_X' and 'Arc_Y' not found in 'UpdatedPath'. No changes made.")
