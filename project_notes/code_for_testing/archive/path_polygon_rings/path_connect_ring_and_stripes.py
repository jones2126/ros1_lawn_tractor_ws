#!/usr/bin/env python

'''
Connect the last waypoint of the ring path to the first point of the stripes.

$ python3 /home/tractor/ros1_lawn_tractor_ws/project_notes/code_for_testing/archive/path_polygon_rings/path_connect_ring_and_stripes.py
'''

import pandas as pd
from openpyxl import load_workbook

print("Starting the process...")

# Define the input file path and sheet names
input_file_path = '/home/tractor/ros1_lawn_tractor_ws/project_notes/paths/Collins_Dr_62/site1_20240513/collins_dr_62_A_from_rosbag_step1_20240513_2.xlsx'
input_sheet_name = 'outer_rings_path'
output_sheet_name = 'coverage_path_refrmttd'

# Read the 'outer_rings_path' sheet
outer_rings_df = pd.read_excel(input_file_path, sheet_name=input_sheet_name, header=None)

# Get the x and y values from the last row
last_x = outer_rings_df.iloc[-1, 0]
last_y = outer_rings_df.iloc[-1, 1]

print(f"Extracted last point from '{input_sheet_name}': ({last_x}, {last_y})")

# Read the 'coverage_path_refrmttd' sheet
coverage_path_df = pd.read_excel(input_file_path, sheet_name=output_sheet_name, header=None)

# Create a new row with the last x and y values
new_row = pd.DataFrame([[last_x, last_y]], columns=[0, 1])

# Concatenate the new row with the existing data
updated_coverage_path_df = pd.concat([new_row, coverage_path_df]).reset_index(drop=True)

# Round all x and y values to 2 decimal places
updated_coverage_path_df = updated_coverage_path_df.round(2)

print(f"Added new point to the top of sheet '{output_sheet_name}' and rounded all values to 2 decimal places")

# Load the existing workbook
book = load_workbook(input_file_path)

# Remove the existing 'coverage_path_refrmttd' sheet
if output_sheet_name in book.sheetnames:
    book.remove(book[output_sheet_name])

# Save and close the workbook to apply the removal
book.save(input_file_path)
book.close()

print(f"Removed existing '{output_sheet_name}' sheet from workbook")

# Write the updated DataFrame back to the 'coverage_path_refrmttd' sheet
with pd.ExcelWriter(input_file_path, engine='openpyxl', mode='a') as writer:
    updated_coverage_path_df.to_excel(writer, sheet_name=output_sheet_name, index=False, header=False)

print(f"Successfully updated '{output_sheet_name}' sheet with new point at the top")
print("Process completed.")