#!/usr/bin/env python

'''
Script that reads the pose_x and pose_y data from a .xlsx file that represent the bounds around an obstacle and calculates the center 
point and radius. This will be used as input to additional path planning scripts.

$ python3 /home/tractor/ros1_lawn_tractor_ws/project_notes/code_for_testing/archive/path_polygon_rings/path_calc_circle_radius_and_center6.py
'''

print(f"Starting the script path_calc_circle_radius_and_center6.py...")

import numpy as np
import pandas as pd
from scipy.optimize import least_squares
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def read_data_from_xlsx(filename, sheet_name, reference_tag):
    # Read data using pandas
    data = pd.read_excel(filename, engine='openpyxl', sheet_name=sheet_name)
    print("Columns in SiteSurvey sheet:", data.columns)

    # Adjust these column names based on the actual names in your XLSX file
    reference_col = 'Reference 1'
    pose_x_col = 'pose_x'
    pose_y_col = 'pose_y'

    # Filter data based on the reference_tag
    filtered_data = data[data[reference_col] == reference_tag]
    
    x_data = filtered_data[pose_x_col].tolist()
    y_data = filtered_data[pose_y_col].tolist()
    
    return x_data, y_data

def save_to_xlsx(filename, sheet_name, reference_tag, center, radius):
    try:
        book = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        book = openpyxl.Workbook()

    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        book.remove(sheet)

    sheet = book.create_sheet(sheet_name)

    headers = ["Reference", "X", "Y", "Radius"]
    sheet.append(headers)

    formatted_center_x = round(center[0], 2)
    formatted_center_y = round(center[1], 2)
    formatted_radius = round(radius, 2)

    print(f"Saving values: Reference={reference_tag}, X={formatted_center_x}, Y={formatted_center_y}, Radius={formatted_radius}")

    sheet.append([reference_tag, formatted_center_x, formatted_center_y, formatted_radius])

    book.save(filename)

def read_obstacle_data_from_xlsx(filename, sheet_name):
    data = pd.read_excel(filename, engine='openpyxl', sheet_name=sheet_name)
    print("Columns in Obstacle sheet:", data.columns)

    data.columns = ['Reference', 'X', 'Y', 'Radius']
    reference = data['Reference'].values
    x = data['X'].values
    y = data['Y'].values
    radius = data['Radius'].values

    print("Extracted References:", reference)
    print("Extracted X coordinates:", x)
    print("Extracted Y coordinates:", y)
    print("Extracted Radii:", radius)

    obstacle_data = list(zip(reference, x, y, radius))
    
    return obstacle_data

def calculate_circle_arc(circle_center, radius, num_points):
    segments = []
    angles = np.linspace(0, 2 * np.pi, num_points + 1)
    for i in range(num_points):
        start_angle = angles[i]
        end_angle = angles[i + 1]
        start_point = (circle_center[0] + radius * np.cos(start_angle), circle_center[1] + radius * np.sin(start_angle))
        end_point = (circle_center[0] + radius * np.cos(end_angle), circle_center[1] + radius * np.sin(end_angle))
        segments.append((start_point, end_point))
    return segments

def save_segments_to_xlsx(filename, reference_tag, segments):
    try:
        book = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        book = openpyxl.Workbook()

    if reference_tag in book.sheetnames:
        sheet = book[reference_tag]
        book.remove(sheet)

    sheet = book.create_sheet(reference_tag)

    headers = ["Start X", "Start Y", "End X", "End Y"]
    sheet.append(headers)

    for start_point, end_point in segments:
        start_x, start_y = round(start_point[0], 2), round(start_point[1], 2)
        end_x, end_y = round(end_point[0], 2), round(end_point[1], 2)
        sheet.append([start_x, start_y, end_x, end_y])

    book.save(filename)

def circle_fit(X, Y):
    x_m = np.mean(X)
    y_m = np.mean(Y)

    def calc_R(xc, yc):
        return np.sqrt((X - xc)**2 + (Y - yc)**2)

    def f_2(c):
        Ri = calc_R(*c)
        return Ri - Ri.mean()

    center_estimate = x_m, y_m
    center_ier = least_squares(f_2, center_estimate)
    xc, yc = center_ier.x

    Ri = calc_R(xc, yc)
    R = Ri.mean()  # The mean of the distances to the center is the radius

    return xc, yc, R

def plot_data_and_circle(X, Y, xc, yc, R, R_inflated, segments=None):
    best_fit_circle = plt.Circle((xc, yc), R, color='blue', fill=False, label='Best Fit Circle')
    inflated_circle = plt.Circle((xc, yc), R_inflated, color='orange', fill=False, linestyle='--', label='Inflated Circle')
    
    fig, ax = plt.subplots()
    ax.add_artist(best_fit_circle)
    ax.add_artist(inflated_circle)
    ax.scatter(X, Y, color='red', label='Data Points')
    
    if segments:
        for start_point, end_point in segments:
            line = plt.Line2D((start_point[0], end_point[0]), (start_point[1], end_point[1]), color='green')
            ax.add_line(line)
    
    ax.set_aspect('equal', adjustable='datalim')
    plt.title('Circle Fit to Data with Inflated Circle')
    plt.xlabel('X')
    plt.ylabel('Y')
    plt.legend()
    plt.grid()
    plt.show()

# File paths and sheet names
folder_path = '/home/tractor/ros1_lawn_tractor_ws/project_notes/paths/Collins_Dr_62/site1_20240513/'
xlsx_filename = folder_path + 'collins_dr_62_A_from_rosbag_step1_20240513_2.xlsx'
tag_reference = 'Obstacle 1'
result_sheet_name = 'Obstcl_list'  # change to Obstcl_list from Obstacle,  It is too close to the segment data sheet name
data_sheet_name = 'SiteSurvey'
inflation = 0.3  # 30% inflation around radius for safety

# Read data from the XLSX file
x_data, y_data = read_data_from_xlsx(xlsx_filename, data_sheet_name, tag_reference)

# Perform circle fitting
xc, yc, R = circle_fit(x_data, y_data)

# Apply inflation to the radius
R_inflated = R * (1 + inflation)

# Save the center and inflated radius data to the XLSX file
save_to_xlsx(xlsx_filename, result_sheet_name, tag_reference, (xc, yc), R_inflated)
print(f'Data saved to sheet {result_sheet_name} in {xlsx_filename}')
print(f"Circle Center: ({xc}, {yc}), Inflated Radius: {R_inflated}")

# Read obstacle data from the 'Obstacle' sheet
obstacle_data = read_obstacle_data_from_xlsx(xlsx_filename, result_sheet_name)
print(f"Obstacle data: {obstacle_data}")

# Calculate and save line segments for each obstacle
num_points = 20
all_segments = []
for reference, x, y, radius in obstacle_data:
    print("in loop to build segments")
    print(reference, x, y, radius)
    segments = calculate_circle_arc((x, y), radius, num_points)
    print(f"segments {segments}")
    save_segments_to_xlsx(xlsx_filename, reference, segments)
    all_segments.extend(segments)

# Plot the data, the best fit circle, the inflated circle, and the line segments
plot_data_and_circle(x_data, y_data, xc, yc, R, R_inflated, segments=all_segments)
