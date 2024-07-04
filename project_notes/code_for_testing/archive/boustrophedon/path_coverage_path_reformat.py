#!/usr/bin/env python

'''
Read the data from the specified sheet, reformat it, and save the results to a new sheet named coverage_path_refrmttd in the same Excel file.
Each row of the input file consists of 4 (x,y) pairs.  Reformating it converts each row (i.e. 4 pairs) into 4 rows with 1 (x,y) pair each. 
This is in support of creating a Boustrophedon coverage path that uses Dubins (i.e. key hole) u-turns.

$ python3 path_coverage_path_reformat.py
'''
import openpyxl
import pandas as pd
import numpy as np

def reformat_excel(input_file_path, input_sheet_name, output_sheet_name):
    # Load the workbook and read the input sheet into a DataFrame
    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook[input_sheet_name]

    # Read the sheet into a DataFrame
    data = sheet.values
    df = pd.DataFrame(data, columns=[
        'Point_1_x', 'Point_1_y', 
        'Point_2_x', 'Point_2_y', 
        'Point_3_x', 'Point_3_y', 
        'Point_4_x', 'Point_4_y'
    ])

    reformatted_data = []

    # Iterate over each row and extract each set of x and y coordinates in the specified order
    for index, row in df.iterrows():
        reformatted_data.append([row['Point_1_x'], row['Point_1_y']])
        reformatted_data.append([row['Point_2_x'], row['Point_2_y']])
        reformatted_data.append([row['Point_3_x'], row['Point_3_y']])
        reformatted_data.append([row['Point_4_x'], row['Point_4_y']])

    # Convert the list of lists into a DataFrame
    reformatted_df = pd.DataFrame(reformatted_data, columns=['X', 'Y'])

    # Write the results to the specified output sheet
    if output_sheet_name in workbook.sheetnames:
        del workbook[output_sheet_name]
    output_sheet = workbook.create_sheet(output_sheet_name)

    # Write the DataFrame to the sheet
    for row in reformatted_df.itertuples(index=False):
        output_sheet.append(row)

    workbook.save(input_file_path)

input_file_path = '/home/tractor/ros1_lawn_tractor_ws/project_notes/paths/Collins_Dr_62/site1_20240513/collins_dr_62_A_from_rosbag_step1_20240513_2.xlsx'
input_sheet_name = 'coverage_path'
output_sheet_name = 'coverage_path_refrmttd'

# Call the function to reformat the Excel sheet
reformat_excel(input_file_path, input_sheet_name, output_sheet_name)
print(f"List of x, y points ready for angle calculations. Use sheet '{output_sheet_name}' in the Excel file '{input_file_path}' as input.")
