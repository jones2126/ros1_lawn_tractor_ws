#!/usr/bin/env python
'''
Read the x, y data for the path, calculate the angle and add the lookahead and speed data.

$ python3 ~/ros1_lawn_tractor_ws/project_notes/code_for_testing/archive/boustrophedon/path_add_angle_to_xy2.py
'''

import openpyxl
import pandas as pd
import math

def update_df_with_angle_lookahead_speed(input_file_path, input_sheet_name, lookahead, speed, output_sheet_name):
    def calculate_angle(x1, y1, x2, y2):
        '''Calculates the directional angle with respect to the positive X-axis.'''
        angle = math.atan2(y2 - y1, x2 - x1)
        if angle < 0:
            angle += 2 * math.pi
        rounded_angle = round(angle, 2)
        print(x1, y1, x2, y2, rounded_angle)
        return rounded_angle


    print("Calculating angles")
    df = pd.read_excel(input_file_path, sheet_name=input_sheet_name, header=None)
    df.columns = ['X', 'Y']

    # Calculate angles for records 0-3
    first_record = calculate_angle(df.iloc[0]['X'], df.iloc[0]['Y'], df.iloc[1]['X'], df.iloc[1]['Y'])
    pair_1 = calculate_angle(df.iloc[1]['X'], df.iloc[1]['Y'], df.iloc[2]['X'], df.iloc[2]['Y'])
    pair_2 = calculate_angle(df.iloc[3]['X'], df.iloc[3]['Y'], df.iloc[4]['X'], df.iloc[4]['Y'])

    # Initialize the angles list
    angles = [first_record]
    
    # Populate the angles list
    for i in range(1, len(df)):
        if i % 4 == 1 or i % 4 == 2:
            angles.append(pair_1)
        elif i % 4 == 3 or i % 4 == 0:
            angles.append(pair_2)

    # Ensure that the length of angles matches the DataFrame
    while len(angles) < len(df):
        angles.append(angles[-1])

    print(f"Length of angles list {len(angles)}, length of DataFrame {len(df)}")

    print("Adding lookahead and speed")
    df['angle'] = angles
    df['lookahead'] = lookahead
    df['speed'] = speed

    print("Writing to Excel file")
    workbook = openpyxl.load_workbook(input_file_path)
    if output_sheet_name in workbook.sheetnames:
        del workbook[output_sheet_name]
    output_sheet = workbook.create_sheet(output_sheet_name)
    for row in df.itertuples(index=False, name=None):
        output_sheet.append(row)
    workbook.save(input_file_path)
    print(f"Data published to sheet: {output_sheet_name} in file: {input_file_path}")

input_file_path = '/home/tractor/ros1_lawn_tractor_ws/project_notes/paths/Collins_Dr_62/site1_20240513/collins_dr_62_A_from_rosbag_step1_20240513_2.xlsx'
input_sheet_name = 'coverage_path_refrmttd'
output_sheet_name = 'path_with_angle'
lookahead = 2.5
speed = 0.75

update_df_with_angle_lookahead_speed(input_file_path, input_sheet_name, lookahead, speed, output_sheet_name)