#!/usr/bin/env python

'''
Script that uses a Boustrophedon approach to building a coverage path.

$ python3 /home/tractor/ros1_lawn_tractor_ws/project_notes/code_for_testing/archive/boustrophedon/path_boustrophedon_coverage_variable_degree2.py

'''
import pandas as pd
import numpy as np
from shapely.geometry import Polygon, LineString, MultiLineString
from shapely.affinity import rotate
from matplotlib import pyplot as plt, patheffects
import os
import csv
from openpyxl import load_workbook

script_name = os.path.basename(__file__)
print(f"running script: {script_name}")

# Adding this line for Jupyter Notebook so the plot function at the end appears in the Jupyter notebook
# Check if the script is being run in a Jupyter notebook.  If so run '%matplotlib inline'
try:
    from IPython import get_ipython
    ipython = get_ipython()
    if ipython is not None:
        ipython.run_line_magic('matplotlib', 'inline')
except (ImportError, NameError):
    pass

def boustrophedon_path_corrected(polygon, line_spacing, angle_degrees):
    # Normalize the angle to be within -90 to 90 degrees for intersection purposes
    angle_degrees = angle_degrees % 180  # Limit the angle within the 0-180 degree range
    if angle_degrees > 90:
        angle_degrees -= 180  # Convert 91-180 degrees to -89 to -1 degrees

    # Convert the angle from degrees to radians
    angle_radians = np.radians(angle_degrees)
    
    # Rotate the polygon to align the slicing lines with the y-axis
    rotated_polygon = rotate(polygon, -angle_degrees, origin='centroid', use_radians=False)
    minx, miny, maxx, maxy = rotated_polygon.bounds

    # Calculate the length needed for lines to cover the entire area after rotation
    bounding_width = maxx - minx
    bounding_height = maxy - miny
    diag_length = np.sqrt(bounding_width**2 + bounding_height**2)

    # Determine the extended bounds for the lines to ensure they cover the entire rotated polygon area
    extended_bounds = diag_length * np.sqrt(2)
    extended_minx = minx - extended_bounds
    extended_maxx = maxx + extended_bounds
    extended_miny = miny - extended_bounds
    extended_maxy = maxy + extended_bounds

    # Calculate the number of lines needed, considering the line spacing
    num_lines = int(np.ceil(extended_bounds * 2 / (line_spacing * np.cos(angle_radians))))

    path = []
    # Generate lines from one side of the bounding box to the other
    for i in range(num_lines + 1):
        # Calculate the start point of the line
        startx = extended_minx + (line_spacing * np.cos(angle_radians)) * i
        starty = extended_miny
        endx = startx
        endy = extended_maxy
        
        # Create a line that will be rotated to the specified angle
        line = LineString([(startx, starty), (endx, endy)])
        # Rotate the line to the specified angle
        rotated_line = rotate(line, angle_degrees, origin='centroid', use_radians=False)
        # Find the intersection of the rotated line with the original polygon
        intersection = polygon.intersection(rotated_line)
        if not intersection.is_empty:
            # Only consider LineStrings or MultiLineString (ignore Points)
            if isinstance(intersection, LineString):
                path.append(intersection)
            elif isinstance(intersection, MultiLineString):
                for line in intersection.geoms:
                    path.append(line)                    

    # Count the total number of line segments excluding the joining segments
    return path, len(path)

# The plot_path function remains the same, now we call it plot_path_corrected
def plot_path(polygon, path_lines, degree):
    fig, ax = plt.subplots()
    x, y = polygon.exterior.xy
    ax.plot(x, y, color='blue', alpha=0.7, linewidth=3, solid_capstyle='round', zorder=2)

    for idx, line in enumerate(path_lines):
        x, y = line.xy
        ax.plot(x, y, color='red', linewidth=2, alpha=0.7, zorder=2)
        # Position the text in the center of the line segment
        text_x = (x[0] + x[1]) / 2
        text_y = (y[0] + y[1]) / 2
        ax.text(text_x, text_y, str(idx + 1), fontsize=8, verticalalignment='center',
                horizontalalignment='center', color='white', path_effects=[
                patheffects.withStroke(linewidth=2, foreground="black")])

    ax.set_title(f'Boustrophedon Path at {degree}-degree')
    plt.xlabel('X coordinate')
    plt.ylabel('Y coordinate')
    plt.grid(True)
    plt.axis('equal')
    # Get the absolute path of the current script
    filepath = os.path.abspath(__file__)
    plt.figtext(0.5, 0.01, f'Reference: {filepath}', ha='center', fontsize=8, color='gray')
    plt.show()

def write_path_to_excel(path, xlsx_file_path, sheet_name):
    # Load the existing workbook
    workbook = load_workbook(xlsx_file_path)
    # Add a new sheet
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]  # Delete the sheet if it already exists to avoid duplicates
    worksheet = workbook.create_sheet(sheet_name)

    # Write the header
    worksheet.append(['x1', 'y1', 'x2', 'y2'])
    
    # Iterate over the LineStrings in the path
    for linestring in path:
        # For each line, write the start and end coordinates
        start_point, end_point = linestring.coords[:]
        worksheet.append([start_point[0], start_point[1], end_point[0], end_point[1]])

    # Save the workbook
    workbook.save(xlsx_file_path)

def read_inner_ring(xlsx_file_path):
    # Read the specified sheet from the Excel file
    df = pd.read_excel(xlsx_file_path, sheet_name='UpdatedPath')
    # Filter the data to include only points with Path_Index = 0
    filtered_df = df[df['Path_Index'] == 0]
    # Create a list of points (tuples) from the filtered data
    polygon_points = list(zip(filtered_df['X'], filtered_df['Y']))
    # Return the polygon created from these points
    return Polygon(polygon_points)

# The main function now correctly handles the angle normalization
def main(angle_degrees):
    # Path to the Excel file
    xlsx_file_path = '/home/tractor/ros1_lawn_tractor_ws/project_notes/paths/Collins_Dr_62/site1_20240513/collins_dr_62_A_from_rosbag_step1_20240513_2.xlsx'  
    print("reading the file: ", xlsx_file_path)
    
    # Read the polygon data from the Excel file
    polygon = read_inner_ring(xlsx_file_path)

    # Generate the Boustrophedon path with the corrected angle and line spacing
    path, num_segments = boustrophedon_path_corrected(polygon, line_spacing=1.8, angle_degrees=angle_degrees)
    print(f"Total line segments: {num_segments}")

    # Plot the path
    plot_path(polygon, path, angle_degrees)
    return path, xlsx_file_path

# Run the main function with a corrected angle
path, xlsx_file_path = main(19)
# The output will be written to a new sheet in the Excel file
write_path_to_excel(path, xlsx_file_path, 'boustrphdn_segmnts')

print("First few lines of boustrophedon line segments")
for point in path[:5]:
    print(point)
print('eoj')
