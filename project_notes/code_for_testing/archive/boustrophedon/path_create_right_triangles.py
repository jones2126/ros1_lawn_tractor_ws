#!/usr/bin/env python

'''
Script that calculates a right triangle points based for line segments from a Boustrophedon coverage path in order to create input for a Dubins path

$ python3 /home/tractor/ros1_lawn_tractor_ws/project_notes/code_for_testing/archive/boustrophedon/path_create_right_triangles.py

'''
import openpyxl
import pandas as pd
import numpy as np

def calculate_x3_y3_solutions(x1, y1, x2, y2, length=0.9):
    """
    Calculate two possible coordinates of the third point (x3, y3) to form a right triangle
    with the line segment from (x1, y1) to (x3, y3) as the hypotenuse. This function returns
    both solutions, allowing for a choice where x3 can be greater than x2.
    """
    dx = x2 - x1
    dy = y2 - y1

    # Normalize the direction vector (dx, dy)
    norm = np.sqrt(dx**2 + dy**2)
    dx /= norm
    dy /= norm

    # Rotate 90 degrees in both directions and scale by the desired length
    dx1, dy1 = -dy * length, dx * length  # First solution
    dx2, dy2 = dy * length, -dx * length  # Second solution

    # Calculate (x3, y3) for both solutions
    x3_1 = x2 + dx1
    y3_1 = y2 + dy1
    x3_2 = x2 + dx2
    y3_2 = y2 + dy2

    return (x3_1, y3_1), (x3_2, y3_2)

def choose_x3_y3_solution(x2, y2, x3_1, y3_1, x3_2, y3_2, preference='greater'):
    """
    Choose the (x3, y3) point based on the preference.

    Parameters:
    - x3_1, y3_1, x2, y2, x3_2, y3_2: Coordinates of x3 and y3 from the first and second solutions, and x2, y2.
    - preference: Can be 'greater' or 'less' to choose which 'side' the hypotenuse should be on in relation to the x-axis.

    Returns:
    - x3, y3: The chosen coordinates of x3 and y3.
    - solution: A string indicating which solution was chosen.
    """
    if preference == 'greater':
        if x3_1 > x2:
            return x3_1, y3_1, "larger on the x-axis"
        else:
            return x3_2, y3_2, "smaller on the x-axis"
    else:  # preference is 'less'
        if x3_1 < x2:
            return x3_1, y3_1, "larger on the x-axis"
        else:
            return x3_2, y3_2, "smaller on the x-axis"

# Define the Excel file path and sheet names
xlsx_file_path = '/home/tractor/ros1_lawn_tractor_ws/project_notes/paths/Collins_Dr_62/site1_20240513/collins_dr_62_A_from_rosbag_step1_20240513_2.xlsx'
input_sheet_name = 'boustrphdn_trimmed'
output_sheet_name = 'coverage_path'

# Load the workbook and read the input sheet into a DataFrame
workbook = openpyxl.load_workbook(xlsx_file_path)
sheet = workbook[input_sheet_name]

# Read the sheet into a DataFrame
data = sheet.values
columns = next(data)[0:4]  # Assuming first row is the header
df = pd.DataFrame(data, columns=['x1', 'y1', 'x2', 'y2'])

# Initialize an empty DataFrame for storing results
results_df = pd.DataFrame(columns=[
    'Point_1_x', 'Point_1_y', 
    'Point_2_x', 'Point_2_y',
    'Point_3_x', 'Point_3_y', 
    'Point_4_x', 'Point_4_y'
])

# Iterate over each line segment in the DataFrame
for index, row in df.iterrows():
    bottom_x, bottom_y, top_x, top_y = row[['x1', 'y1', 'x2', 'y2']]

    # Handle the bottom of the line segment
    (hyptns_x_opt_a, hyptns_y_opt_a), (hyptns_x_opt_b, hyptns_y_opt_b) = calculate_x3_y3_solutions(top_x, top_y, bottom_x, bottom_y)
    pt1_hyptns_x, pt1_hyptns_y, solution = choose_x3_y3_solution(bottom_x, bottom_y, hyptns_x_opt_a, hyptns_y_opt_a, hyptns_x_opt_b, hyptns_y_opt_b, preference='less')

    # Handle the top of the line segment
    (hyptns_x_opt_a, hyptns_y_opt_a), (hyptns_x_opt_b, hyptns_y_opt_b) = calculate_x3_y3_solutions(bottom_x, bottom_y, top_x, top_y)
    pt2_hyptns_x, pt2_hyptns_y, solution = choose_x3_y3_solution(top_x, top_y, hyptns_x_opt_a, hyptns_y_opt_a, hyptns_x_opt_b, hyptns_y_opt_b, preference='greater')

    # Store points in the DataFrame
    results_df.loc[index] = [pt1_hyptns_x, pt1_hyptns_y, bottom_x, bottom_y, top_x, top_y, pt2_hyptns_x, pt2_hyptns_y]

# Write the results to the specified output sheet
if output_sheet_name in workbook.sheetnames:
    del workbook[output_sheet_name]
output_sheet = workbook.create_sheet(output_sheet_name)

# Write the DataFrame to the sheet
for row in results_df.itertuples(index=False):
    output_sheet.append(row)

workbook.save(xlsx_file_path)

print(f"Results saved to sheet: {output_sheet_name} in file: {xlsx_file_path}")
