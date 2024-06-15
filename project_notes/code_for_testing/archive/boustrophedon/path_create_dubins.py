#!/usr/bin/env python

'''
Read the data from the specified sheet, and create Dubins path u-turns.

$ python3 /home/tractor/ros1_lawn_tractor_ws/project_notes/code_for_testing/archive/boustrophedon/path_create_dubins.py
'''
import openpyxl
import pandas as pd
import math
import dubins

def generate_path(x0, y0, x1, y1, theta0, theta1, turning_radius, step_size):
    q0 = (x0, y0, theta0)
    q1 = (x1, y1, theta1)
    continuous_path = dubins.shortest_path(q0, q1, turning_radius)
    calculated_path_segment, _ = continuous_path.sample_many(step_size) # takes the continuous Dubins path and samples it at intervals specified by step_size
    return calculated_path_segment

def calculate_curvature(p1, p2, p3):
    x1, y1 = p1[0], p1[1]
    x2, y2 = p2[0], p2[1]
    x3, y3 = p3[0], p3[1]

    # Calculate the semi-perimeter
    a = math.sqrt((x1 - x2)**2 + (y1 - y2)**2)
    b = math.sqrt((x2 - x3)**2 + (y2 - y3)**2)
    c = math.sqrt((x3 - x1)**2 + (y3 - y1)**2)
    s = (a + b + c) / 2

    # Check for collinearity or near-collinearity
    area = s * (s - a) * (s - b) * (s - c)
    if area <= 0:
        return 0  # Curvature is zero if the points are collinear

    # Calculate the radius of the circumscribed circle
    radius = a * b * c / (4 * math.sqrt(area))

    # Calculate the curvature
    curvature = 1 / radius if radius != 0 else 0

    return curvature

def calculate_curvature_and_output_excel(drive_path, output_file_path, output_sheet_name, alternate_lookahead, curvature_threshold):
    workbook = openpyxl.load_workbook(output_file_path)
    
    # Check if the sheet already exists and delete it if it does
    if output_sheet_name in workbook.sheetnames:
        del workbook[output_sheet_name]

    # Create a new sheet
    output_sheet = workbook.create_sheet(output_sheet_name)
    
    # Write the header
    output_sheet.append(['x', 'y', 'curvature', 'lookahead'])
    
    for i in range(1, len(drive_path) - 1):
        p1, p2, p3 = drive_path[i - 1], drive_path[i], drive_path[i + 1]
        curvature = calculate_curvature(p1, p2, p3)

        # Update lookahead distance based on curvature
        if curvature > curvature_threshold:
            lookahead = alternate_lookahead
        else:
            lookahead = p2[3]  # Use existing lookahead if curvature is not above threshold

        # Update the tuple in drive_path
        drive_path[i] = (p2[0], p2[1], p2[2], lookahead, p2[4])

        # Write data to the sheet
        output_sheet.append([p2[0], p2[1], curvature, lookahead])
    
    workbook.save(output_file_path)
    print(f"Data published to sheet: {output_sheet_name} in file: {output_file_path}")
    return drive_path

def main():
    def read_waypoints(input_file_path, input_sheet_name):
        df = pd.read_excel(input_file_path, sheet_name=input_sheet_name, header=None)
        df.columns = ['x', 'y', 'theta', 'lookahead', 'speed']
        return df.values.tolist()

    def generate_drive_path(waypoints, turning_radius, step_size, lookahead, speed):
        drive_path = []
        x0, y0, theta0 = 0.0, 0.0, 0.0
        for idx, point in enumerate(waypoints):
            x1, y1, theta1 = point[0], point[1], point[2]
            if idx == 0:
                x0, y0, theta0 = x1, y1, theta1   # For the first line, just set the initial x, y, theta
                continue
            # Generate path from the previous to the current waypoint
            calculated_path_segment = generate_path(x0, y0, x1, y1, theta0, theta1, turning_radius, step_size)
            for segment in calculated_path_segment:
                segment += (lookahead, speed)   # Appending lookahead and speed to each point in the path
                drive_path.append(segment)
            x0, y0, theta0 = x1, y1, theta1     # Updating the start point for the next segment
        return drive_path

    def write_output_excel(output_file_path, output_sheet_name, drive_path):
        workbook = openpyxl.load_workbook(output_file_path)
        
        # Check if the sheet already exists and delete it if it does
        if output_sheet_name in workbook.sheetnames:
            del workbook[output_sheet_name]

        # Create a new sheet
        output_sheet = workbook.create_sheet(output_sheet_name)

        # Write the header
        output_sheet.append(['x', 'y', 'theta', 'lookahead', 'speed'])
        
        for p in drive_path:
            output_sheet.append(p)

        workbook.save(output_file_path)
        return len(drive_path)

    # main routine
    turning_radius = 1.3
    step_size = 0.3
    lookahead = 2.5
    alternate_lookahead = 1.0
    curvature_threshold = 0.02
    speed = 0.75

    working_directory = "/home/tractor/ros1_lawn_tractor_ws/project_notes/paths/Collins_Dr_62/"
    input_file_path = working_directory + "site1_20240513/collins_dr_62_A_from_rosbag_step1_20240513_2.xlsx"
    input_sheet_name = 'path_with_angle'
    output_sheet_name = 'path_dubins'
    #output_file_curvature_of_waypoints = working_directory + "test_PID_curvature_of_waypoints.xlsx"

    waypoints = read_waypoints(input_file_path, input_sheet_name)
    drive_path = generate_drive_path(waypoints, turning_radius, step_size, lookahead, speed)    # Process 1: Path Generation
    print("done with generate_drive_path")

    # Call the function to calculate curvature and output to Excel
    drive_path = calculate_curvature_and_output_excel(drive_path, input_file_path, output_sheet_name, alternate_lookahead, curvature_threshold)
    print("done with calculate_curvature_and_output_excel")

    total_count = write_output_excel(input_file_path, output_sheet_name, drive_path)            # Process 2: Output File Creation and Data Appending
    print(f"Output file: {input_file_path}")
    print("Record count:", total_count, "turning_radius:", turning_radius, "step_size:", step_size)
    print("eoj")

if __name__ == "__main__":
    main()
