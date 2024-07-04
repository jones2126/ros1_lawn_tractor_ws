# #!/usr/bin/env python

# '''
# Read the x, y data for the path and flip (i.e. reverse) the sequence.

# $ python3 /home/tractor/ros1_lawn_tractor_ws/project_notes/code_for_testing/archive/boustrophedon/path_reverse_dubins.py
# '''
# print("Starting the reverse process")
# import pandas as pd

# # Define the input file path
# input_file_path = '/home/tractor/ros1_lawn_tractor_ws/project_notes/paths/Collins_Dr_62/site1_20240513/collins_dr_62_A_from_rosbag_step1_20240513_2.xlsx'
# sheet_name='path_dubins'

# # Load the Excel file
# excel_data = pd.ExcelFile(input_file_path)

# # Load the 'path_dubins' sheet into a DataFrame with a one-row header
# path_dubins_df = pd.read_excel(input_file_path, sheet_name, header=0)

# # Reverse the order of the DataFrame
# reversed_path_dubins_df = path_dubins_df.iloc[::-1].reset_index(drop=True)

# # Define the output file path (same as input, we just add a new sheet)
# output_file_path = input_file_path

# # Save the reversed DataFrame into a new sheet in the same Excel file
# with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
#     reversed_path_dubins_df.to_excel(writer, sheet_name='path_dubins_reversed', index=False, header=True)

# print(f"Reversed path saved to sheet 'path_dubins_reversed' in the file: {output_file_path}")

#!/usr/bin/env python
'''
Read the x, y data for the path and flip (i.e. reverse) the sequence.
'''
#!/usr/bin/env python
'''
Read the x, y data for the path and flip (i.e. reverse) the sequence.
'''
print("Starting the reverse process")
import pandas as pd
from openpyxl import load_workbook

# Define the input file path
input_file_path = '/home/tractor/ros1_lawn_tractor_ws/project_notes/paths/Collins_Dr_62/site1_20240513/collins_dr_62_A_from_rosbag_step1_20240513_2.xlsx'
input_sheet_name = 'coverage_path_refrmttd'

# Load the 'coverage_path_refrmttd' sheet into a DataFrame without header
try:
    path_df = pd.read_excel(input_file_path, sheet_name=input_sheet_name, header=None, engine='openpyxl')
except Exception as e:
    print(f"Error reading Excel file: {e}")
    exit(1)

# Reverse the order of the DataFrame
reversed_path_df = path_df.iloc[::-1].reset_index(drop=True)

# Load the existing workbook
try:
    book = load_workbook(input_file_path)
except Exception as e:
    print(f"Error loading workbook: {e}")
    exit(1)

# Remove the existing sheet if it exists
if input_sheet_name in book.sheetnames:
    book.remove(book[input_sheet_name])
    
# Save and close the workbook
book.save(input_file_path)
book.close()

# Save the reversed DataFrame into the same sheet in the Excel file
try:
    with pd.ExcelWriter(input_file_path, engine='openpyxl', mode='a') as writer:
        reversed_path_df.to_excel(writer, sheet_name=input_sheet_name, index=False, header=False)
except Exception as e:
    print(f"Error writing to Excel file: {e}")
    exit(1)

print(f"Reversed path saved to sheet '{input_sheet_name}' in the file: {input_file_path}")