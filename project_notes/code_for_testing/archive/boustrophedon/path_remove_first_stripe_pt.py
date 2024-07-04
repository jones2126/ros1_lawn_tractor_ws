#!/usr/bin/env python
'''
Delete the first record from the specified Excel sheet.

$ python3 ~/ros1_lawn_tractor_ws/project_notes/code_for_testing/archive/boustrophedon/path_remove_first_stripe.py
'''
import pandas as pd
from openpyxl import load_workbook

print("Starting the process...")

# Define the input file path and sheet name
input_file_path = '/home/tractor/ros1_lawn_tractor_ws/project_notes/paths/Collins_Dr_62/site1_20240513/collins_dr_62_A_from_rosbag_step1_20240513_2.xlsx'
input_sheet_name = 'coverage_path_refrmttd'

# Read the specified sheet
df = pd.read_excel(input_file_path, sheet_name=input_sheet_name, header=None)

print(f"Read {len(df)} rows from '{input_sheet_name}' sheet; Now deleting first record")

# Delete the first record
df = df.iloc[1:].reset_index(drop=True)

print(f"Deleted the first record. {len(df)} rows remaining; Now removing sheet from workbook.")

# Load the existing workbook
book = load_workbook(input_file_path)

# Remove the existing sheet
if input_sheet_name in book.sheetnames:
    book.remove(book[input_sheet_name])

# Save and close the workbook to apply the removal
book.save(input_file_path)
book.close()

print(f"Removed existing '{input_sheet_name}' sheet from workbook; Now writing new sheet.")

# Write the updated DataFrame back to the sheet
with pd.ExcelWriter(input_file_path, engine='openpyxl', mode='a') as writer:
    df.to_excel(writer, sheet_name=input_sheet_name, index=False, header=False)

print(f"Successfully updated '{input_sheet_name}' sheet with the first record removed")
print("Process completed.")