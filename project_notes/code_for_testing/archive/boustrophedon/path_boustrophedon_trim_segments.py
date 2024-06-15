#!/usr/bin/env python

'''
Script that uses a Boustrophedon approach to building a coverage path.

$ python3 /home/tractor/ros1_lawn_tractor_ws/project_notes/code_for_testing/archive/boustrophedon/path_boustrophedon_trim_segments.py

'''
import openpyxl

# Define number of lines to remove from beginning and end
beginning_lines_to_remove = 2
end_lines_to_remove = 2

# Define input and output file paths
xlsx_file_path = '/home/tractor/ros1_lawn_tractor_ws/project_notes/paths/Collins_Dr_62/site1_20240513/collins_dr_62_A_from_rosbag_step1_20240513_2.xlsx'
input_sheet_name = 'boustrphdn_segmnts'
output_sheet_name = 'boustrphdn_trimmed'

# Load the workbook and select the input sheet
print(f"Reading sheet: {input_sheet_name} from file: {xlsx_file_path}")
workbook = openpyxl.load_workbook(xlsx_file_path)
input_sheet = workbook[input_sheet_name]

# Read data from the input sheet
data = list(input_sheet.iter_rows(values_only=True))

# Print total rows before trim
print(f"Total rows before trim: {len(data)}")

# Remove lines from the beginning and end
trimmed_data = data[beginning_lines_to_remove:len(data) - end_lines_to_remove]

# Print total rows after trim
print(f"Total rows after trim: {len(trimmed_data)}")

# Create a new sheet or replace if it already exists
if output_sheet_name in workbook.sheetnames:
    workbook.remove(workbook[output_sheet_name])
output_sheet = workbook.create_sheet(output_sheet_name)

# Write the trimmed data to the new sheet
for row in trimmed_data:
    output_sheet.append(row)

# Save the workbook
workbook.save(xlsx_file_path)

# Print confirmation message
print(f"Trimmed data saved to sheet: {output_sheet_name} in file: {xlsx_file_path}")
